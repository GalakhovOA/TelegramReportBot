# main.py
import os
import asyncio
from io import BytesIO
from datetime import datetime
from dotenv import load_dotenv
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, BotCommand, InputFile
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes

import config
import database
import json

# load .env
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DOTENV_PATH = os.path.join(BASE_DIR, '.env')
if os.path.exists(DOTENV_PATH):
    load_dotenv(dotenv_path=DOTENV_PATH)
else:
    load_dotenv()

TOKEN = os.getenv('BOT_TOKEN')
if not TOKEN:
    print("ERROR: BOT_TOKEN not found in env (BOT_TOKEN)")

user_states = {}

def safe_state(uid):
    st = user_states.get(uid)
    if not st:
        st = {'mode': 'idle', 'step': 0, 'data': {}, 'editing': False}
        user_states[uid] = st
    return st

def build_main_menu():
    kb = [
        [InlineKeyboardButton("Отчет МКК", callback_data='role_mkk')],
        [InlineKeyboardButton("Отчеты РТП", callback_data='role_rtp')],
        [InlineKeyboardButton("Отчеты РМ/МН", callback_data='role_rm')],
        [InlineKeyboardButton("Сменить ФИ/РТП", callback_data='change_info')]
    ]
    return InlineKeyboardMarkup(kb)

# --- Helpers for xlsx generation (used by RM) ---
def generate_xlsx_for_report(title: str, rows: list, columns: list):
    try:
        import openpyxl
    except Exception:
        raise
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]
    for c_idx, (_, col_title) in enumerate(columns, start=1):
        ws.cell(row=1, column=c_idx, value=col_title)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (col_key, _) in enumerate(columns, start=1):
            value = row.get(col_key, "") if isinstance(row, dict) else ""
            ws.cell(row=r_idx, column=c_idx, value=value)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# --- Handlers ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message or update.effective_message
    await msg.reply_text("Выберите роль:", reply_markup=build_main_menu())

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not query:
        return
    await query.answer()
    uid = query.from_user.id
    data = query.data or ""
    st = user_states.get(uid, {})

    # return to main
    if data == 'return_to_menu':
        user_states.pop(uid, None)
        await query.edit_message_text("Выберите роль:", reply_markup=build_main_menu())
        return

    # role selection (common)
    if data.startswith('role_'):
        role = data.split('_',1)[1]
        # For RTP and RM: require password if user not verified
        if role in ('rtp', 'rm'):
            if database.is_user_verified(uid):
                user_states[uid] = {'mode': role, 'step': 0, 'data': {}, 'editing': False}
                await handle_role_selection(query, uid, role)
                return
            else:
                user_states[uid] = {'mode': 'awaiting_password_for', 'await_role': role}
                try:
                    await query.edit_message_text("Введите пароль для доступа в раздел руководителя:")
                except Exception:
                    await query.message.reply_text("Введите пароль для доступа в раздел руководителя:")
                return
        else:
            # mkk flow
            user_states[uid] = {'mode': role, 'step': 0, 'data': {}, 'editing': False}
            await handle_role_selection(query, uid, role)
            return

    # change_info
    if data == 'change_info':
        user_states[uid] = {'mode': 'change_fi_enter_name'}
        try:
            await query.edit_message_text("Введите ваше ФИ (как хотите, чтобы оно сохранялось):")
        except Exception:
            await query.message.reply_text("Введите ваше ФИ (как хотите, чтобы оно сохранялось):")
        return

    # choose_rtp_{idx}
    if data.startswith('choose_rtp_'):
        try:
            idx = int(data.split('_')[2])
        except Exception:
            await query.edit_message_text("Ошибка выбора. Попробуйте снова.")
            return
        if idx < 0 or idx >= len(config.RTP_LIST):
            await query.edit_message_text("Некорректный индекс РТП.")
            return
        selected = config.RTP_LIST[idx]
        # if in change_flow (user entered new name earlier)
        if st.get('change_flow'):
            new_name = st.get('new_name')
            if not new_name:
                await query.edit_message_text("Ошибка: имя не найдено в состоянии.")
                return
            database.add_user(uid, 'mkk', new_name, selected)
            user_states.pop(uid, None)
            await query.edit_message_text(f"Готово. Ваше имя '{new_name}' привязано к РТП: {selected}.")
            return

        role = st.get('mode', 'idle')
        if role == 'rtp':
            # user choosing their own FI as RTP
            database.add_user(uid, 'rtp', selected)
            # when RTP chooses own FI, ensure verified flag set (they passed password earlier)
            database.set_user_verified(uid, 1)
            user_states[uid] = {'mode': 'rtp', 'step': 0, 'data': {}, 'editing': False}
            await query.edit_message_text(f"Вы вошли как РТП: {selected}")
            await show_manager_menu(query)
            return

        # registration flow for MKK
        name = st.get('name')
        if name:
            database.add_user(uid, 'mkk', name, selected)
            st.pop('choosing_rtp', None); st.pop('name', None)
            st.update({'step': 0, 'data': {}, 'editing': False, 'mode': 'mkk'})
            await query.edit_message_text(f"Привязка к {selected} успешна. Начинаем отчёт.")
            await ask_next_question(query.message, uid)
            return

        await query.edit_message_text("Непонятный контекст выбора РТП.")
        return

    # choose_rm_{idx} - RM selects their FI from list
    if data.startswith('choose_rm_'):
        try:
            idx = int(data.split('_')[2])
        except Exception:
            await query.edit_message_text("Ошибка выбора РМ/МН.")
            return
        if idx < 0 or idx >= len(config.RM_MN_LIST):
            await query.edit_message_text("Некорректный индекс.")
            return
        chosen = config.RM_MN_LIST[idx]
        # register user as rm and mark verified
        database.add_user(uid, 'rm', chosen)
        database.set_user_verified(uid, 1)
        user_states[uid] = {'mode': 'rm', 'step': 0, 'data': {}, 'editing': False}
        kb = [
            [InlineKeyboardButton("Список РТП", callback_data='rm_show_rtps')],
            [InlineKeyboardButton("Вернуться в меню", callback_data='return_to_menu')]
        ]
        await query.edit_message_text(f"Вы вошли как РМ/МН: {chosen}", reply_markup=InlineKeyboardMarkup(kb))
        return

    # role_rm -> show RM menu (entry)
    if data == 'role_rm':
        if database.is_user_verified(uid):
            await handle_role_selection(query, uid, 'rm')
            return
        else:
            user_states[uid] = {'mode': 'awaiting_password_for', 'await_role': 'rm'}
            try:
                await query.edit_message_text("Введите пароль для доступа в раздел руководителя:")
            except Exception:
                await query.message.reply_text("Введите пароль для доступа в раздел руководителя:")
            return

    # RM menu interactions
    if data == 'rm_show_rtps':
        date = datetime.now().strftime('%Y-%m-%d')
        sent_status = database.get_rtp_combined_status_for_all(config.RTP_LIST, date)
        kb = []
        for i, fi in enumerate(config.RTP_LIST):
            status = "✅" if sent_status.get(fi, False) else "❌"
            kb.append([InlineKeyboardButton(f"{fi} {status}", callback_data=f"rm_choose_rtp_{i}")])
        kb.append([InlineKeyboardButton("Объединить все РТП (глобально) и скачать", callback_data='rm_combine_all')])
        kb.append([InlineKeyboardButton("Вернуться в меню", callback_data='return_to_menu')])
        await query.edit_message_text("Список РТП (статус отправки объединённого отчёта):", reply_markup=InlineKeyboardMarkup(kb))
        return

    if data.startswith('rm_choose_rtp_'):
        # format: rm_choose_rtp_{i}
        try:
            idx = int(data.split('_')[3])
        except Exception:
            await query.edit_message_text("Ошибка выбора.")
            return
        if idx < 0 or idx >= len(config.RTP_LIST):
            await query.edit_message_text("Некорректный индекс.")
            return
        chosen = config.RTP_LIST[idx]
        date = datetime.now().strftime('%Y-%m-%d')
        combined = database.get_rtp_combined(chosen, date)
        if not combined:
            await query.edit_message_text(f"РТП {chosen} не отправлял объединённый отчёт на {date}.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Назад", callback_data='rm_show_rtps')]]))
            return
        text = f"Объединённый отчёт РТП {chosen} на {date}:\n\n{config.format_report(combined)}"
        kb = [
            [InlineKeyboardButton("📥 Скачать .xlsx", callback_data=f"download_rtp_{idx}")],
            [InlineKeyboardButton("Назад", callback_data='rm_show_rtps')]
        ]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb))
        return

    if data == 'rm_combine_all':
        date = datetime.now().strftime('%Y-%m-%d')
        all_combined = database.get_all_rtp_combined_on_date(date)
        if not all_combined:
            await query.edit_message_text(f"Нет объединённых отчётов от РТП на {date}.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Назад", callback_data='rm_show_rtps')]]))
            return
        aggregated = {}
        fckp_products = []
        for rtp_fi, rdata in all_combined:
            for k, v in rdata.items():
                if k == 'fckp_products' and isinstance(v, list):
                    fckp_products.extend(v)
                else:
                    try:
                        aggregated[k] = aggregated.get(k, 0) + float(v or 0)
                    except Exception:
                        pass
        aggregated['fckp_products'] = fckp_products
        aggregated['fckp_realized'] = len(fckp_products)
        text = f"Глобальный объединённый отчёт за {date}:\n\n{config.format_report(aggregated)}"
        kb = [
            [InlineKeyboardButton("📥 Скачать глобальный .xlsx", callback_data="download_global")],
            [InlineKeyboardButton("Назад", callback_data='rm_show_rtps')]
        ]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb))
        return

    if data.startswith('download_rtp_'):
        try:
            idx = int(data.split('_')[2])
        except Exception:
            await query.edit_message_text("Ошибка скачивания.")
            return
        if idx < 0 or idx >= len(config.RTP_LIST):
            await query.edit_message_text("Некорректный индекс.")
            return
        rtp_fi = config.RTP_LIST[idx]
        date = datetime.now().strftime('%Y-%m-%d')
        rdata = database.get_rtp_combined(rtp_fi, date)
        if not rdata:
            await query.edit_message_text("Отчёт не найден.")
            return
        rows = []
        for q in config.QUESTIONS:
            rows.append({'key': q['question'], 'value': rdata.get(q['key'], 0)})
        prod_counts = {}
        for p in rdata.get('fckp_products', []):
            prod_counts[p] = prod_counts.get(p, 0) + 1
        for prod in config.FCKP_OPTIONS:
            rows.append({'key': prod, 'value': prod_counts.get(prod, 0)})
        cols = [('key', 'Поле'), ('value', 'Значение')]
        try:
            bio = generate_xlsx_for_report(f"{rtp_fi}_{date}", rows, cols)
            filename = f"rtp_{rtp_fi.replace(' ','_')}_{date}.xlsx"
            await context.bot.send_document(chat_id=uid, document=InputFile(bio, filename=filename))
        except Exception as e:
            await query.edit_message_text(f"Ошибка формирования файла: {e}")
        return

    if data == 'download_global':
        date = datetime.now().strftime('%Y-%m-%d')
        all_combined = database.get_all_rtp_combined_on_date(date)
        rows = []
        for rtp_fi, rdata in all_combined:
            row = {'rtp': rtp_fi}
            for q in config.QUESTIONS:
                row[q['key']] = rdata.get(q['key'], 0)
            row['fckp_count'] = len(rdata.get('fckp_products', []))
            rows.append(row)
        cols = [('rtp', 'RTP')]
        for q in config.QUESTIONS:
            cols.append((q['key'], q['question']))
        cols.append(('fckp_count', 'FCKP count'))
        try:
            bio = generate_xlsx_for_report(f"global_{date}", rows, cols)
            filename = f"global_combined_{date}.xlsx"
            await context.bot.send_document(chat_id=uid, document=InputFile(bio, filename=filename))
        except Exception as e:
            await query.edit_message_text(f"Ошибка формирования файла: {e}")
        return

    # role_rtp menu (entry)
    if data == 'role_rtp':
        if database.is_user_verified(uid):
            kb = [[InlineKeyboardButton(fi, callback_data=f"choose_rtp_{i}")] for i, fi in enumerate(config.RTP_LIST)]
            kb.append([InlineKeyboardButton("Вернуться в меню", callback_data='return_to_menu')])
            await query.edit_message_text("Выберите ваше ФИ (РТП):", reply_markup=InlineKeyboardMarkup(kb))
            return
        else:
            user_states[uid] = {'mode': 'awaiting_password_for', 'await_role': 'rtp'}
            try:
                await query.edit_message_text("Введите пароль для доступа в раздел РТП:")
            except Exception:
                await query.message.reply_text("Введите пароль для доступа в раздел РТП:")
            return

    # RTP manager actions
    if data == 'rtp_menu':
        await show_manager_menu(query)
        return

    if data == 'rtp_show_reports':
        date = datetime.now().strftime('%Y-%m-%d')
        manager_fi = database.get_user_name(uid)
        employees = database.get_employees(manager_fi)
        reports = database.get_all_reports_on_date(date, manager_fi)
        reported_ids = [u for u,_ in reports]
        text = f"Отчеты на {date}:\n"
        for u_id, name in employees:
            status = '✅' if u_id in reported_ids else '❌'
            text += f"Сотрудник {name or str(u_id)}: {status}\n"
        kb = [[InlineKeyboardButton("Детальный отчет на дату", callback_data='rtp_detailed_reports')], [InlineKeyboardButton("Назад", callback_data='rtp_menu')]]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb))
        return

    if data == 'rtp_detailed_reports':
        date = datetime.now().strftime('%Y-%m-%d')
        manager_fi = database.get_user_name(uid)
        reports = database.get_all_reports_on_date(date, manager_fi)
        text = f"Детальные отчеты на {date}:\n\n"
        for u_id, rdata in reports:
            name = database.get_user_name(u_id) or str(u_id)
            text += f"Сотрудник {name}:\n{config.format_report(rdata)}\n\n"
        kb = [[InlineKeyboardButton("Вернуться в меню", callback_data='rtp_menu')]]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb))
        return

    if data == 'rtp_combine_reports':
        date = datetime.now().strftime('%Y-%m-%d')
        manager_fi = database.get_user_name(uid)
        reports = database.get_all_reports_on_date(date, manager_fi)
        if not reports:
            await query.edit_message_text("Нет отчетов на сегодня.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Назад", callback_data='rtp_menu')]]))
            return
        combined = {}
        fckp_products = []
        for _, r in reports:
            for k, v in r.items():
                if k == 'fckp_products' and isinstance(v, list):
                    fckp_products.extend(v)
                else:
                    try:
                        combined[k] = combined.get(k, 0) + float(v or 0)
                    except Exception:
                        pass
        combined['fckp_products'] = fckp_products
        combined['fckp_realized'] = len(fckp_products)
        text = f"Объединённый отчёт на {date}:\n\n{config.format_report(combined)}\n\n" + config.OPERATIONAL_DEFECTS_BLOCK
        kb = [
            [InlineKeyboardButton("Отправить РМ/МН", callback_data='rtp_send_to_rm')],
            [InlineKeyboardButton("Назад", callback_data='rtp_menu')]
        ]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb))
        return

    if data == 'rtp_send_to_rm':
        manager_fi = database.get_user_name(uid)
        date = datetime.now().strftime('%Y-%m-%d')
        reports = database.get_all_reports_on_date(date, manager_fi)
        if not reports:
            await query.edit_message_text("Нет отчетов для объединения/отправки.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Назад", callback_data='rtp_menu')]]))
            return
        combined = {}
        fckp_products = []
        for _, r in reports:
            for k, v in r.items():
                if k == 'fckp_products' and isinstance(v, list):
                    fckp_products.extend(v)
                else:
                    try:
                        combined[k] = combined.get(k, 0) + float(v or 0)
                    except Exception:
                        pass
        combined['fckp_products'] = fckp_products
        combined['fckp_realized'] = len(fckp_products)
        database.save_rtp_combined(manager_fi, combined, date)
        await query.edit_message_text("Объединённый отчёт сохранён и доступен РМ/МН.")
        return

    # FCKP product picking
    if data.startswith('fckp_prod_'):
        prod = data.split('fckp_prod_',1)[1]
        st = safe_state(uid)
        st.setdefault('fckp_products', [])
        st['fckp_products'].append(prod)
        st['fckp_left'] = st.get('fckp_left', 0) - 1
        left = st.get('fckp_left', 0)
        if left > 0:
            kb = [[InlineKeyboardButton(p, callback_data=f"fckp_prod_{p}")] for p in config.FCKP_OPTIONS]
            try:
                await query.edit_message_text(f"Вы выбрали {prod}. Осталось указать ещё {left} ФЦКП.", reply_markup=InlineKeyboardMarkup(kb))
            except Exception:
                pass
            return
        else:
            st['data']['fckp_products'] = st.get('fckp_products', [])
            st['data']['fckp_realized'] = len(st.get('fckp_products', []))
            try:
                await query.edit_message_text("Все ФЦКП указаны ✅")
            except Exception:
                pass
            st['step'] = st.get('step',0) + 1
            await ask_next_question(query.message, uid)
            return

    # download individual user report (RTP view)
    if data.startswith('download_user_'):
        try:
            parts = data.split('_')
            target_uid = int(parts[2])
        except Exception:
            await query.edit_message_text("Ошибка скачивания.")
            return
        date = datetime.now().strftime('%Y-%m-%d')
        rpt = database.get_report(target_uid, date)
        if not rpt:
            await query.edit_message_text("Отчёт не найден.")
            return
        rows = []
        for q in config.QUESTIONS:
            rows.append({'key': q['question'], 'value': rpt.get(q['key'], 0)})
        prod_counts = {}
        for p in rpt.get('fckp_products', []):
            prod_counts[p] = prod_counts.get(p,0) + 1
        for prod in config.FCKP_OPTIONS:
            rows.append({'key': prod, 'value': prod_counts.get(prod,0)})
        cols = [('key','Поле'), ('value','Значение')]
        try:
            bio = generate_xlsx_for_report(f"user_{target_uid}_{date}", rows, cols)
            filename = f"user_{target_uid}_{date}.xlsx"
            await context.bot.send_document(chat_id=uid, document=InputFile(bio, filename=filename))
        except Exception as e:
            await query.edit_message_text(f"Ошибка формирования файла: {e}")
        return

    if data == 'send_report':
        success, msg_text = await send_personal_report_to_manager(uid, context)
        try:
            if success:
                await query.edit_message_text("Отчёт успешно отправлен руководителю.")
            else:
                await query.edit_message_text(f"Отчет отправлен, но {msg_text}")
        except Exception:
            try:
                await query.message.reply_text("Отчёт отправлен (или произошла ошибка, проверьте лог).")
            except Exception:
                pass
        return

    # fallback
    return

# role selection helper
async def handle_role_selection(query_or_message, user_id, role):
    name = database.get_user_name(user_id)
    if role == 'rtp':
        kb = [[InlineKeyboardButton(fi, callback_data=f"choose_rtp_{i}")] for i,fi in enumerate(config.RTP_LIST)]
        kb.append([InlineKeyboardButton("Вернуться в меню", callback_data='return_to_menu')])
        try:
            await query_or_message.edit_message_text("Выберите ваше ФИ (РТП):", reply_markup=InlineKeyboardMarkup(kb))
        except Exception:
            try:
                await query_or_message.reply_text("Выберите ваше ФИ (РТП):", reply_markup=InlineKeyboardMarkup(kb))
            except Exception:
                pass
        return

    if role == 'rm':
        kb = [[InlineKeyboardButton(fi, callback_data=f"choose_rm_{i}")] for i, fi in enumerate(config.RM_MN_LIST)]
        kb.append([InlineKeyboardButton("Вернуться в меню", callback_data='return_to_menu')])
        try:
            await query_or_message.edit_message_text("Выберите ваше ФИ (РМ/МН):", reply_markup=InlineKeyboardMarkup(kb))
        except Exception:
            try:
                await query_or_message.reply_text("Выберите ваше ФИ (РМ/МН):", reply_markup=InlineKeyboardMarkup(kb))
            except Exception:
                pass
        return

    # MKK flow: ask for name then choose RТП
    if role == 'mkk':
        if name:
            manager_fi = database.get_manager_fi_for_employee(user_id)
            if manager_fi:
                user_states[user_id] = {'mode': role, 'step': 0, 'data': {}, 'editing': False}
                try:
                    await query_or_message.edit_message_text("Роль выбрана. Начинаем заполнение отчёта.")
                except Exception:
                    try:
                        await query_or_message.reply_text("Роль выбрана. Начинаем заполнение отчёта.")
                    except Exception:
                        pass
                await start_filling(query_or_message, user_id)
                return
            else:
                user_states[user_id] = {'mode': role, 'choosing_rtp': True, 'name': name}
                await show_rtp_buttons(query_or_message, "Выберите вашего РТП:")
                return
        else:
            user_states[user_id] = {'mode': role, 'entering_name': True}
            try:
                await query_or_message.edit_message_text("Пожалуйста, введите ваше имя (ФИ):")
            except Exception:
                try:
                    await query_or_message.reply_text("Пожалуйста, введите ваше имя (ФИ):")
                except Exception:
                    pass
            return

async def show_rtp_buttons(query_or_message, text):
    kb = [[InlineKeyboardButton(fi, callback_data=f"choose_rtp_{i}")] for i,fi in enumerate(config.RTP_LIST)]
    kb.append([InlineKeyboardButton("Вернуться в меню", callback_data='return_to_menu')])
    try:
        await query_or_message.reply_text(text, reply_markup=InlineKeyboardMarkup(kb))
    except Exception:
        try:
            await query_or_message.message.reply_text(text, reply_markup=InlineKeyboardMarkup(kb))
        except Exception:
            pass

async def show_manager_menu(q):
    kb = [
        [InlineKeyboardButton("Показать отчеты на дату", callback_data='rtp_show_reports')],
        [InlineKeyboardButton("Детальный отчет на дату", callback_data='rtp_detailed_reports')],
        [InlineKeyboardButton("Объединить и показать отчеты на дату", callback_data='rtp_combine_reports')],
        [InlineKeyboardButton("Вернуться в меню", callback_data='return_to_menu')]
    ]
    try:
        await q.edit_message_text("Меню руководителя:", reply_markup=InlineKeyboardMarkup(kb))
    except Exception:
        try:
            await q.message.reply_text("Меню руководителя:", reply_markup=InlineKeyboardMarkup(kb))
        except Exception:
            pass

# messages handler
async def message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    if not msg:
        return
    uid = msg.from_user.id
    text = (msg.text or "").strip()
    st = user_states.get(uid, {})

    if not st:
        if text.lower() == "вернуться в меню":
            await start(update, context)
            return
        await msg.reply_text("Сессия не запущена. Нажмите /start.")
        return

    if text.lower() == "вернуться в меню":
        user_states.pop(uid, None)
        await start(update, context)
        return

    # Password entry flow for RTP/RM (variant B)
    if st.get('mode') == 'awaiting_password_for':
        await_role = st.get('await_role')
        if text.lower() == 'отмена' or text.lower() == 'cancel':
            user_states.pop(uid, None)
            await msg.reply_text("Отмена. Возврат в меню.", reply_markup=build_main_menu())
            return
        # check password
        if text == config.ADMIN_PASSWORD:
            # ensure user row exists and mark verified
            database.add_user(uid, await_role)
            # don't override name; set_user_name called only if present
            database.set_user_verified(uid, 1)
            user_states[uid] = {'mode': await_role, 'step': 0, 'data': {}, 'editing': False}
            await msg.reply_text("Пароль верный. Доступ предоставлен.")
            await handle_role_selection(msg, uid, await_role)
            return
        else:
            await msg.reply_text("Неверный пароль. Попробуйте снова")
            return

    # change FI flow
    if st.get('mode') == 'change_fi_enter_name':
        entered_name = text
        st['new_name'] = entered_name
        st['change_flow'] = True
        await show_rtp_buttons(msg, f"Вы ввели имя: {entered_name}\nТеперь выберите вашего РТП из списка:")
        return

    # Registration flows (MKK name entering)
    if st.get('entering_name'):
        name = text
        role = st.get('mode','idle')
        st['name'] = name
        st.pop('entering_name', None)
        database.add_user(uid, 'mkk' if role == 'mkk' else role, name)
        if role == 'mkk':
            st['choosing_rtp'] = True
            await show_rtp_buttons(update, "Выберите вашего РТП:")
        else:
            await msg.reply_text("Выберите ваше ФИ из списка кнопок.")
        return

    if st.get('choosing_rtp'):
        await msg.reply_text("Пожалуйста, выберите РТП из списка кнопок.")
        return

    # Now questionnaire: accept floats (and ints)
    if 'step' not in st:
        return

    step = st['step']
    if step < len(config.QUESTIONS):
        q = config.QUESTIONS[step]
        # Accept float-like input (allow comma)
        t = text.replace(',', '.')
        try:
            if t == '':
                val = 0.0
            else:
                val = float(t)
        except Exception:
            await msg.reply_text("Пожалуйста, введите число (можно дробное, например 1.5 либо 0,7).")
            return

        # special handling for fckp_realized (asks for product choices)
        if q['key'] == 'fckp_realized':
            n = int(val)
            st['data'][q['key']] = n
            if n > 0:
                st['fckp_left'] = n
                st['fckp_products'] = []
                kb = [[InlineKeyboardButton(p, callback_data=f"fckp_prod_{p}")] for p in config.FCKP_OPTIONS]
                await msg.reply_text(f"Вы указали {n} ФЦКП. Выберите оформленный продукт (1/{n}):", reply_markup=InlineKeyboardMarkup(kb))
                return
            else:
                st['step'] += 1
                await ask_next_question(msg, uid)
                return
        else:
            st['data'][q['key']] = str(val)
            st['step'] += 1
            await ask_next_question(msg, uid)
            return
    else:
        await msg.reply_text("Опрос завершён. Для возврата в меню нажмите 'Вернуться в меню' или /start.")
        return

async def ask_next_question(msgobj, uid):
    st = safe_state(uid)
    step = st.get('step', 0)
    if step < len(config.QUESTIONS):
        q = config.QUESTIONS[step]
        current = st.get('data', {}).get(q['key'], '')
        try:
            await msgobj.reply_text(f"{q['question']} {f'(текущее: {current})' if current != '' else ''}")
        except Exception:
            try:
                await msgobj.message.reply_text(f"{q['question']} {f'(текущее: {current})' if current != '' else ''}")
            except Exception:
                pass
    else:
        await finish_report(msgobj, uid)

async def start_filling(query_or_message, uid, editing=False):
    st = safe_state(uid)
    st['editing'] = editing
    st['step'] = 0
    if not editing:
        st['data'] = {}
    try:
        await query_or_message.edit_message_text("Начинаем заполнение отчёта.")
    except Exception:
        try:
            await query_or_message.reply_text("Начинаем заполнение отчёта.")
        except Exception:
            pass
    await ask_next_question(query_or_message, uid)

async def finish_report(msgobj, uid):
    st = safe_state(uid)
    data = st.get('data', {}) or {}
    if 'fckp_products' in st and st.get('fckp_products'):
        data['fckp_products'] = st.get('fckp_products')
        data['fckp_realized'] = len(st.get('fckp_products'))
    # ensure all questions present
    for q in config.QUESTIONS:
        data.setdefault(q['key'], 0)
    try:
        if st.get('mode') != 'idle':
            database.save_report(uid, data)
    except Exception as e:
        print("DB save_report error:", e)
    formatted = config.format_report(data)
    try:
        await msgobj.reply_text(f"Итоговый отчет:\n{formatted}")
    except Exception:
        try:
            await msgobj.message.reply_text(f"Итоговый отчет:\n{formatted}")
        except Exception:
            pass
    # Final actions menu (no change_info here per request)
    kb = [
        [InlineKeyboardButton("Редактировать", callback_data='edit_report')]
    ]
    if st.get('mode') == 'mkk':
        kb[0].insert(1, InlineKeyboardButton("Отправить руководителю", callback_data='send_report'))
    try:
        await msgobj.reply_text("Действия:", reply_markup=InlineKeyboardMarkup(kb))
    except Exception:
        pass

async def send_personal_report_to_manager(uid, context):
    date = datetime.now().strftime('%Y-%m-%d')
    rpt = database.get_report(uid, date)
    if not rpt:
        return False, "Отчёт не найден"
    formatted = config.format_report(rpt)
    name = database.get_user_name(uid) or str(uid)
    manager_fi = database.get_manager_fi_for_employee(uid)
    if not manager_fi:
        return False, "Руководитель не привязан"
    manager_id = database.get_manager_id_by_fi(manager_fi)
    if not manager_id:
        return False, f"руководитель {manager_fi} не найден в системе"
    try:
        await context.bot.send_message(chat_id=manager_id, text=f"Отчёт от сотрудника {name} на {date}:\n{formatted}")
        return True, "Отчёт отправлен"
    except Exception as e:
        print("send_personal_report_to_manager error:", e)
        return False, "Ошибка отправки"

async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    print("Error:", context.error)

async def set_commands(app):
    try:
        await app.bot.set_my_commands([BotCommand("start", "Начать работу с ботом")])
    except Exception as e:
        print("set_commands error:", e)

if __name__ == '__main__':
    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler('start', start))
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, message_handler))
    app.add_error_handler(error_handler)
    asyncio.get_event_loop().run_until_complete(set_commands(app))
    print("Bot started")
    app.run_polling()
